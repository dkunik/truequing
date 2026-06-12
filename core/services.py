from openpyxl import load_workbook
import base64
import json
import zlib

from django.db import transaction

from core.models import AlbumUsuario, Figurita, MovimientoAlbum
from core.types import EstadoCanje


def inicializar_album_usuario(usuario, coleccion):
    figuritas = Figurita.objects.filter(coleccion=coleccion)

    for figurita in figuritas:
        AlbumUsuario.objects.get_or_create(
            usuario=usuario,
            figurita=figurita,
            defaults={"cantidad": 0},
        )


@transaction.atomic
def registrar_movimiento(
    usuario,
    figurita,
    delta,
    tipo="manual",
    descripcion="",
):
    item, _ = AlbumUsuario.objects.get_or_create(
        usuario=usuario,
        figurita=figurita,
        defaults={"cantidad": 0},
    )

    nueva_cantidad = item.cantidad + delta

    if nueva_cantidad < 0:
        raise ValueError("La cantidad no puede quedar negativa.")

    item.cantidad = nueva_cantidad
    item.save()

    MovimientoAlbum.objects.create(
        usuario=usuario,
        figurita=figurita,
        delta=delta,
        tipo=tipo,
        descripcion=descripcion,
    )

    return item


def agregar_figurita(usuario, figurita, tipo="manual", descripcion=""):
    return registrar_movimiento(
        usuario=usuario,
        figurita=figurita,
        delta=1,
        tipo=tipo,
        descripcion=descripcion,
    )


def quitar_figurita(usuario, figurita, tipo="manual", descripcion=""):
    return registrar_movimiento(
        usuario=usuario,
        figurita=figurita,
        delta=-1,
        tipo=tipo,
        descripcion=descripcion,
    )


def obtener_estado_canje_usuario(usuario, coleccion):
    repetidas = set(
        AlbumUsuario.objects.filter(
            usuario=usuario,
            figurita__coleccion=coleccion,
            cantidad__gt=1,
        ).values_list("figurita_id", flat=True)
    )

    faltantes = set(
        AlbumUsuario.objects.filter(
            usuario=usuario,
            figurita__coleccion=coleccion,
            cantidad=0,
        ).values_list("figurita_id", flat=True)
    )

    return EstadoCanje(
        usuario=usuario.username,
        coleccion=coleccion.nombre,
        repetidas=repetidas,
        faltantes=faltantes,
    )


def calcular_canje(estado_a, estado_b):
    a_tiene_para_b = estado_a.repetidas & estado_b.faltantes
    b_tiene_para_a = estado_b.repetidas & estado_a.faltantes

    return {
        "usuario_a": estado_a.usuario,
        "usuario_b": estado_b.usuario,
        "a_tiene_para_b": a_tiene_para_b,
        "b_tiene_para_a": b_tiene_para_a,
        "total": len(a_tiene_para_b) + len(b_tiene_para_a),
    }


def exportar_estado_canje_compacto(usuario, coleccion):
    estado = obtener_estado_canje_usuario(usuario, coleccion)

    figuritas = list(
        Figurita.objects
        .filter(coleccion=coleccion)
        .order_by("pais__codigo", "numero")
    )

    indices = {
        figurita.id: i
        for i, figurita in enumerate(figuritas)
    }

    n = len(figuritas)

    repetidas_bits = [0] * n
    faltantes_bits = [0] * n

    for figurita_id in estado.repetidas:
        repetidas_bits[indices[figurita_id]] = 1

    for figurita_id in estado.faltantes:
        faltantes_bits[indices[figurita_id]] = 1

    def bits_a_bytes(bits):
        data = bytearray()

        for i in range(0, len(bits), 8):
            byte = 0

            for bit_index, bit in enumerate(bits[i:i + 8]):
                if bit:
                    byte |= 1 << bit_index

            data.append(byte)

        return bytes(data)

    payload = {
        "v": 1,
        "u": estado.usuario,
        "c": estado.coleccion,
        "n": n,
        "r": base64.urlsafe_b64encode(
            bits_a_bytes(repetidas_bits)
        ).decode("ascii"),
        "f": base64.urlsafe_b64encode(
            bits_a_bytes(faltantes_bits)
        ).decode("ascii"),
    }

    payload_json = json.dumps(
        payload,
        separators=(",", ":"),
        ensure_ascii=False,
    )

    payload_comprimido = zlib.compress(
        payload_json.encode("utf-8")
    )

    return base64.urlsafe_b64encode(
        payload_comprimido
    ).decode("ascii")


def decodificar_estado_canje_compacto(payload_qr, coleccion):
    data = zlib.decompress(
        base64.urlsafe_b64decode(payload_qr.encode("ascii"))
    )

    payload = json.loads(data.decode("utf-8"))

    figuritas = list(
        Figurita.objects
        .filter(coleccion=coleccion)
        .order_by("pais__codigo", "numero")
    )

    def bytes_a_indices(texto_base64):
        raw = base64.urlsafe_b64decode(texto_base64.encode("ascii"))
        indices = []

        for byte_index, byte in enumerate(raw):
            for bit_index in range(8):
                if byte & (1 << bit_index):
                    indices.append(byte_index * 8 + bit_index)

        return indices

    repetidas_indices = bytes_a_indices(payload["r"])
    faltantes_indices = bytes_a_indices(payload["f"])

    repetidas = {
        figuritas[i].id
        for i in repetidas_indices
        if i < len(figuritas)
    }

    faltantes = {
        figuritas[i].id
        for i in faltantes_indices
        if i < len(figuritas)
    }

    return EstadoCanje(
        usuario=payload["u"],
        coleccion=payload["c"],
        repetidas=repetidas,
        faltantes=faltantes,
    )
@transaction.atomic
def importar_album_desde_excel(usuario, coleccion, archivo):
    wb = load_workbook(archivo, data_only=True)

    actualizadas = 0
    errores = []

    paises = {
        pais.nombre: pais
        for pais in coleccion.paises.all()
    }

    for nombre_hoja in wb.sheetnames:
        if nombre_hoja not in paises:
            errores.append(f"Hoja ignorada: {nombre_hoja}")
            continue

        pais = paises[nombre_hoja]
        ws = wb[nombre_hoja]

        for fila in ws.iter_rows(min_row=2, values_only=True):
            numero, cantidad = fila[0], fila[1]

            if numero is None:
                continue

            try:
                numero = int(numero)
                cantidad = int(cantidad or 0)
            except ValueError:
                errores.append(
                    f"{nombre_hoja}: fila con valores inválidos"
                )
                continue

            figurita = Figurita.objects.get(
                coleccion=coleccion,
                pais=pais,
                numero=numero,
            )

            item, _ = AlbumUsuario.objects.get_or_create(
                usuario=usuario,
                figurita=figurita,
                defaults={"cantidad": 0},
            )

            delta = cantidad - item.cantidad

            if delta != 0:
                registrar_movimiento(
                    usuario=usuario,
                    figurita=figurita,
                    delta=delta,
                    tipo="ajuste",
                    descripcion="Importación desde Excel",
                )

                actualizadas += 1

    return {
        "actualizadas": actualizadas,
        "errores": errores,
    }
